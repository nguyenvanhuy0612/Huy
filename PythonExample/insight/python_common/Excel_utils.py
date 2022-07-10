def example1():
    # Reading an excel file using Python
    import xlrd

    # Give the location of the file
    loc = ("WorkspaceCCaaS.xlsx")

    # To open Workbook
    wb = xlrd.open_workbook(loc)
    sheet = wb.sheet_by_index(0)

    # For row 0 and column 0
    print(sheet.cell_value(0, 0))


def excell_xlsx_file():
    # import openpyxl module
    import openpyxl

    # Give the location of the file
    path = "WorkspaceCCaaS.xlsx"

    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)

    # Get workbook active sheet object
    # from the active attribute
    sheet_obj = wb_obj.active

    # Cell objects also have a row, column,
    # and coordinate attributes that provide
    # location information for the cell.

    # Note: The first row or
    # column integer is 1, not 0.

    # Cell object is created by using
    # sheet object's cell() method.
    cell_obj = sheet_obj.cell(row=1, column=1)

    # Print value of cell object
    # using the value attribute
    print(cell_obj.value)

    max_col = sheet_obj.max_column

    # Loop will print all columns name
    for i in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row=3, column=i)
        print(cell_obj.value)


def excel_chart():
    # import openpyxl module
    import openpyxl

    # import BubbleChart, Reference, Series class
    # from openpyxl.chart sub_module
    from openpyxl.chart import BubbleChart, Reference, Series

    # Call a Workbook() function of openpyxl
    # to create a new blank Workbook object
    wb = openpyxl.Workbook()

    # Get workbook active sheet
    # from the active attribute.
    sheet = wb.active

    rows = [
        ("Number of Products", "Sales in USD", "Market share"),
        (14, 12200, 15),
        (20, 60000, 33),
        (18, 24400, 10),
        (22, 32000, 42),
    ]

    # write content of each row in 1st, 2nd and 3rd
    # column of the active sheet respectively.
    for row in rows:
        sheet.append(row)

    # Create object of BubbleChart class
    chart = BubbleChart()

    # create data for plotting
    xvalues = Reference(sheet, min_col=1,
                        min_row=2, max_row=5)

    yvalues = Reference(sheet, min_col=2,
                        min_row=2, max_row=5)

    size = Reference(sheet, min_col=3,
                     min_row=2, max_row=5)

    # create a 1st series of data
    series = Series(values=yvalues, xvalues=xvalues,
                    zvalues=size, title="2013")

    # add series data to the chart object
    chart.series.append(series)

    # set the title of the chart
    chart.title = " BUBBLE-CHART "

    # set the title of the x-axis
    chart.x_axis.title = " X_AXIS "

    # set the title of the y-axis
    chart.y_axis.title = " Y_AXIS "

    # add chart to the sheet
    # the top-left corner of a chart
    # is anchored to cell E2 .
    sheet.add_chart(chart, "E2")

    # save the file
    wb.save("bubbleChart.xlsx")


if __name__ == '__main__':
    excel_chart()
